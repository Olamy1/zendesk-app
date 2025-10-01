from __future__ import annotations  # 👈 optional but helps typing
from datetime import datetime, timedelta
from typing import Dict, Any, List, Tuple, Optional
import io, os
from zoneinfo import ZoneInfo
from dateutil import parser as date_parser
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NY = ZoneInfo("America/New_York")

# Colors (match your script)
COLOR_OVER_30 = "FF3300"
COLOR_OVER_20 = "FF9933"
COLOR_OVER_10 = "FFFF99"
COLOR_UNDER_10 = "C1F0C8"
HEADER_COLOR   = "E8E8E8"
TAB_COLORS = {
    "Ticket Breakdown": "FFFF99",
    "AIMS": "D86DCD",
    "R&A": "83CCEB",
    "Policy": "47D359",
    "Cross-Functional": "FFFF00",
}

TICKET_AGE_COLUMN_NAME = "Ticket Age Status"

def last_wednesday(today: datetime) -> datetime:
    offset = (today.weekday() - 2) % 7  # Wed=2
    return (today - timedelta(days=offset)).replace(hour=0, minute=0, second=0, microsecond=0, tzinfo=NY)

def compute_meeting_window() -> Dict[str, str]:
    now = datetime.now(NY)
    start = last_wednesday(now)
    tuesday_noon = (start + timedelta(days=6)).replace(hour=12, minute=0, second=0, microsecond=0)
    end = min(now, tuesday_noon)
    return {"start": start.isoformat(), "end": end.isoformat()}

def _age_bucket(days_open: int) -> str:
    if days_open > 30: return "Over 30 Days"
    if days_open > 20: return "Over 20 Days"
    if days_open > 10: return "Over 10 Days"
    return "Under 10 Days"

def _parse_iso(ts: str) -> datetime:
    dt = date_parser.parse(ts)
    if dt.tzinfo is None:
        return dt.replace(tzinfo=NY)
    return dt.astimezone(NY)

def build_ticket_rows(
    tickets: List[Dict[str, Any]],
    status_map: Dict[int, Dict[str, str]],
    win: Dict[str, str],
    bucketed: bool = True
) -> List[Dict[str, Any]]:
    """
    Build enriched ticket rows for frontend tables.

    Args:
        tickets: List of Zendesk tickets.
        status_map: Map of ticket_id -> {status, updated_at, resolved_at}.
        win: Meeting window {"start": iso, "end": iso}.
        bucketed: If True, compute age buckets (>30/>20/>10/<10).
                  If False, leave ageBucket=None but still include raw ageDays.

    Returns:
        List of rows ready for frontend consumption.
    """
    start = _parse_iso(win["start"])
    end   = _parse_iso(win["end"])
    now   = datetime.now(NY)
    rows: List[Dict[str, Any]] = []

    for t in tickets:
        tid = t["id"]
        created = _parse_iso(t["created_at"]) if t.get("created_at") else now
        days = (now - created).days
        st  = (t.get("status") or "").lower()

        # Age bucket only if requested
        age_bucket = _age_bucket(days) if bucketed else None

        info = None
        if isinstance(tid, int) or (isinstance(tid, str) and tid.isdigit()):
            info = status_map.get(int(tid))

        # closedByMeeting flag
        cbm = False
        if info and st in {"solved", "closed"}:
            ts = info.get("resolved_at") or info.get("updated_at")
            if ts:
                rdt = _parse_iso(ts)
                cbm = (start <= rdt <= end)

        rows.append({
            "id": tid,
            "subject": t.get("subject", ""),
            "group": t.get("group_id"),
            "status": st,
            "assignee_id": t.get("assignee_id"),
            "assignee_name": None,  # optional enrichment later
            "ageBucket": age_bucket,   # bucketed label if enabled
            "ageDays": days,           # always include raw age
            "closedByMeeting": cbm
        })

    return rows


def _fill_color_for_age(age: str) -> Optional[PatternFill]:
    mp = {
        "Over 30 Days": COLOR_OVER_30,
        "Over 20 Days": COLOR_OVER_20,
        "Over 10 Days": COLOR_OVER_10,
        "Under 10 Days": COLOR_UNDER_10,
    }
    hex_ = mp.get(age)
    return PatternFill(start_color=hex_, end_color=hex_, fill_type="solid") if hex_ else None

def _add_all_borders(ws):
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for c in r:
            if c.value is not None:
                c.border = thin

def _autosize(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 1

def make_ticket_workbook(rows: List[Dict[str, Any]]) -> Tuple[bytes, str]:
    """
    Builds a workbook matching your “Ticket Breakdown” presentation and returns (bytes, filename).
    Now includes both Age Bucket and Age Days (hidden by default).
    """
    today_display = datetime.now(NY).strftime("%-m.%-d.%Y") if os.name != "nt" else datetime.now(NY).strftime("%#m.%#d.%Y")
    filename = f"Ticket Breakdown {today_display}.xlsx"

    wb = openpyxl.Workbook()
    # Sheets in order
    target_order = ['Ticket Breakdown', 'AIMS', 'R&A', 'Policy', 'Cross-Functional']
    for i, name in enumerate(target_order):
        if i == 0:
            ws = wb.active
            ws.title = f"Ticket Breakdown {today_display}"
        else:
            ws = wb.create_sheet(name)
        # Tab colors
        if name in TAB_COLORS:
            ws.sheet_properties.tabColor = TAB_COLORS[name]

        # Updated headers → add Age Days
        headers = [
            "Ticket ID", "Ticket Subject", "Ticket group", "Ticket Status",
            "Ticket created", "Ticket updated", "Ticket Assignee",
            TICKET_AGE_COLUMN_NAME,  # e.g. "Ticket Age Status"
            "Ticket Age (Days)",     # NEW COLUMN
            "Closed by Meeting?"
        ]
        ws.append(headers)

        # Header styling
        header_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
        bold_font = Font(bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = bold_font
            cell.alignment = Alignment(horizontal="center")

        # Write rows
        for r in rows:
            ws.append([
                r["id"], r["subject"], r.get("group"), r["status"],
                "", "", r.get("assignee_name") or r.get("assignee_id"),
                r.get("ageBucket"),    # bucket view
                r.get("ageDays"),      # raw numeric age
                "Yes" if r["closedByMeeting"] else "No"
            ])

        # Color rows by bucket column
        age_col_idx = headers.index(TICKET_AGE_COLUMN_NAME) + 1
        for row_idx in range(2, ws.max_row + 1):
            age = str(ws.cell(row=row_idx, column=age_col_idx).value or "").strip()
            fill = _fill_color_for_age(age)
            if fill:
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = fill

        # Hide Age Days column by default
        age_days_col_idx = headers.index("Ticket Age (Days)") + 1
        ws.column_dimensions[openpyxl.utils.get_column_letter(age_days_col_idx)].hidden = True

        _add_all_borders(ws)
        _autosize(ws)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read(), filename

